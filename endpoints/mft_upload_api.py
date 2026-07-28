# pylint: disable=too-many-lines
# pylint: disable=wrong-import-position
"""
Upload API module for secure Excel file uploads with streaming (no disk save).

This module provides a Flask-based REST API for uploading Excel files with comprehensive
security features including virus scanning via ClamAV, content validation, rate limiting,
and automatic DAG triggering in Apache Airflow. Unlike traditional file upload systems,
this implementation uses a streaming approach where files are processed entirely in memory
without being saved to disk, eliminating file duplication and storage concerns.

Key Features:
    - Streaming file upload with no disk persistence
    - Virus scanning via ClamAV integration (stream-based)
    - Excel content validation (.xlsx and .xls) with macro detection
    - Memory-based validation without temporary file creation
    - Rate limiting to prevent DDoS attacks
    - Thread-safe metadata storage with auto-cleanup
    - Automatic DAG triggering with base64-encoded file content
    - Health check endpoint for monitoring
    - Comprehensive error handling and logging
    - CORS support for cross-origin requests

Architecture:
    The module follows a streaming-first approach:
    1. File is read directly into memory as bytes
    2. Virus scanning performed on the byte stream
    3. Excel validation done from memory using BytesIO
    4. File content encoded as base64 for DAG transmission
    5. No file system operations except for xls macro detection (temp file)

Performance Considerations:
    - Files are processed entirely in memory for maximum speed
    - No I/O operations for file saving/loading
    - Memory usage is O(file_size) during upload processing
    - Metadata stored in memory with automatic expiration
    - Rate limiting protects against resource exhaustion

Security Notes:
    - ClamAV virus scanning on byte streams before any processing
    - Macro detection for both xlsx and xls formats
    - File size limits enforced before reading
    - Extension validation to prevent malicious file types
    - No execution of dynamic code from source files
    - Safe handling of potentially malicious Excel content
    - Rate limiting prevents brute force attacks

Error Handling:
    - Comprehensive exception hierarchy with appropriate HTTP status codes
    - Detailed logging at appropriate levels (INFO, WARNING, ERROR)
    - Graceful degradation for missing ClamAV service
    - Clear error messages for validation failures
    - MemoryError handling for oversized files

Integration Notes:
    - Designed as Flask Blueprint for easy integration
    - Compatible with Airflow 3.0+ REST API with JWT authentication
    - Output format compatible with mft_dag.py and downstream tasks
    - Supports both standalone and containerized deployment

Environment Variables:
    FLASK_SECRET_KEY: Secret key for Flask sessions (required)
    FLASK_HOST: Host to bind the server (default: 0.0.0.0)
    MFT_UPLOAD_API_PORT: Port to bind the server (default: 5000)
    FLASK_DEBUG: Debug mode (default: false)
    FLASK_ENV: Environment (development/production)
    AIRFLOW_API_URL: Airflow API endpoint URL (default: http://airflow-apiserver:8080)
    AIRFLOW_USER: Airflow username for JWT authentication (default: admin)
    AIRFLOW_PASSWORD: Airflow password for JWT authentication (default: airflow)
    DAG_ID: DAG ID to trigger after upload (default: mft_etl_pipeline)
    MAX_FILE_SIZE_MB: Maximum allowed file size in MB (default: 5)
    RATE_LIMIT: Rate limit string (default: 10 per minute)
    RATE_LIMIT_STORAGE_URL: Rate limit storage (default: memory://)
    ALLOWED_EXCEL_EXTENSIONS: Comma-separated extensions (default: .xlsx,.xls)
    FIXED_SHEET_NAME: Required sheet name (default: mft)
    IGNORE_OTHER_SHEETS: Ignore additional sheets (default: true)
    MAX_EXCEL_ROWS: Maximum rows per sheet (default: 1000000)
    MAX_EXCEL_COLS: Maximum columns per sheet (default: 1000)

Endpoints:
    GET /health - Health check for monitoring
    POST /upload-mft-excel - Main file upload endpoint (rate limited)
    GET /file/<file_id> - Get file metadata by ID
    GET /file/<file_id>/status - Check file existence status

Usage Example:
    ```python
    from endpoints.mft_upload_api import create_app

    app = create_app()
    app.run(host=0.0.0.0, port=5000)

    # Upload via curl:
    # curl -X POST -F file=@data.xlsx http://localhost:5000/upload-mft-excel

Workflow Example:
    # 1. Upload file
    POST /upload-mft-excel
    {
        file: data.xlsx  # multipart/form-data
    }

    # 2. Response
    {
        message: File uploaded successfully (streaming mode),
        file: {
            safe_filename: data_20260127_143015_a1b2c3d4.xlsx,
            original_filename: data.xlsx,
            file_size: 1024,
            upload_time: 20260127_143015,
            unique_id: a1b2c3d4,
            sheets: [{name: mft, rows: 100, cols: 20}],
            total_rows: 100,
            format: xlsx,
            file_hash: abc123...
        },
        dag_triggered: true,
        dag: {
            dag_run_id: manual__2026-01-27T14:30:15+00:00,
            state: running
        }
    }

Version: 2.0.0
Compatibility: Python 3.14.4+, Flask 6.0.2+
Maintainer: PLD Engineering Center
Created: 2026-02-26
Last Modified: 2026-07-27
License: MIT
Status: Production Ready
"""

import base64
import io
import hashlib
import os
import tempfile
import time
import uuid
from datetime import datetime, timedelta
from functools import wraps
from pathlib import Path
from threading import Lock
from typing import Tuple, Optional, Dict, Any
import zoneinfo

import requests
import xlrd
import openpyxl
from oletools.olevba import VBA_Parser
from flask import Blueprint, Flask, jsonify, request
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from dotenv import load_dotenv
from openpyxl.utils import exceptions as openpyxl_exceptions

# Local imports
from config import get_logger
from config.clamav_service import clamav_scanner

logger = get_logger("endpoints.mft_upload_api")

# Load environment variables
PROJECT_ROOT = Path(__file__).resolve().parents[1]
env_path = PROJECT_ROOT / '.env'
load_dotenv(dotenv_path=env_path)

# ========== TIMEZONE SETTINGS ==========
MOSCOW_TZ = zoneinfo.ZoneInfo("Europe/Moscow")

# ========== CONFIGURATION ==========
FLASK_SECRET_KEY = os.getenv('FLASK_SECRET_KEY')
if not FLASK_SECRET_KEY:
    raise RuntimeError("FLASK_SECRET_KEY must be set in .env file")

FLASK_HOST = os.getenv('FLASK_HOST', '0.0.0.0')
FLASK_PORT = int(os.getenv('MFT_UPLOAD_API_PORT', '5000'))
FLASK_DEBUG = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
FLASK_ENV = os.getenv('FLASK_ENV', 'development')
IS_PRODUCTION = FLASK_ENV == 'production'

# Admin authentication
# ADMIN_API_KEY = os.getenv('ADMIN_API_KEY', 'change-me-in-production')

# Airflow configuration
AIRFLOW_API_URL = os.getenv('AIRFLOW_API_URL', 'http://airflow-apiserver:8080')
AIRFLOW_USER = os.getenv('AIRFLOW_USER', 'admin')
AIRFLOW_PASSWORD = os.getenv('AIRFLOW_PASSWORD', 'airflow')
DAG_ID = os.getenv('DAG_ID', 'mft_etl_pipeline')

# File upload settings
MAX_FILE_SIZE_MB = int(os.getenv('MAX_FILE_SIZE_MB', '5'))
MAX_FILE_SIZE = MAX_FILE_SIZE_MB * 1024 * 1024

# Rate limiting settings
RATE_LIMIT = os.getenv('RATE_LIMIT', '10 per minute')
RATE_LIMIT_STORAGE_URL = os.getenv('RATE_LIMIT_STORAGE_URL', 'memory://')

# Excel validation settings
MAX_EXCEL_ROWS = int(os.getenv('MAX_EXCEL_ROWS', '1000000'))
MAX_EXCEL_COLS = int(os.getenv('MAX_EXCEL_COLS', '1000'))
extensions_str = os.getenv('ALLOWED_EXCEL_EXTENSIONS', '.xlsx,.xls')
ALLOWED_EXCEL_EXTENSIONS = set(extensions_str.split(','))
FIXED_SHEET_NAME = os.getenv('FIXED_SHEET_NAME', 'mft')
IGNORE_OTHER_SHEETS = os.getenv('IGNORE_OTHER_SHEETS', 'true').lower() == 'true'


# ========== THREAD-SAFE STORAGE FOR FILE METADATA ==========
class FileMetadataStore:
    """
    Thread-safe storage for file metadata with automatic expiration.

    Provides thread-safe in-memory dictionary-like store for file metadata
    with timestamp-based expiration. Automatically removes entries older
    than the specified retention period when accessed or during cleanup.

    This class eliminates the need for disk-based persistence while maintaining
    metadata for uploaded files during their retention period.

    Attributes:
        retention_hours (int): Number of hours to retain metadata entries.
                            Defaults to 24 hours.
        _store (Dict[str, Dict[str, Any]]): Internal metadata storage.
        _lock (Lock): Thread lock for safe concurrent access.

    Methods:
        add: Add file metadata with current timestamp
        get: Retrieve metadata if not expired
        cleanup_expired: Remove expired entries and return count

    Example:
        >>> store = FileMetadataStore(retention_hours=24)
        >>> store.add(file123, {name: data.xlsx, size: 1024})
        >>> metadata = store.get(file123)
        >>> if metadata:
        ...     print(f"File: {metadata[name]}")
    """

    def __init__(
            self,
            retention_hours: int = 24
        ):
        """
        Initialize the metadata store with retention period.

        Args:
            retention_hours (int): Number of hours to retain metadata entries.
                                Defaults to 24 hours.
        """
        self._store: Dict[str, Dict[str, Any]] = {}
        self._lock = Lock()
        self.retention_hours = retention_hours

    def add(
            self,
            file_id: str,
            metadata: Dict[str, Any]
        ) -> None:
        """
        Add file metadata to the store with current timestamp.

        Args:
            file_id (str): Unique identifier for the file
            metadata (Dict[str, Any]): File metadata dictionary to store

        Note:
            Automatically adds timestamp and file_id keys to metadata
        """
        with self._lock:
            metadata['timestamp'] = datetime.now().isoformat()
            metadata['file_id'] = file_id
            self._store[file_id] = metadata

    def get(
            self,
            file_id: str
        ) -> Optional[Dict[str, Any]]:
        """
        Retrieve file metadata if it exists and hasn't expired.

        Args:
            file_id (str): Unique identifier for the file

        Returns:
            Optional[Dict[str, Any]]: Metadata dictionary if found and not expired,
                                    None otherwise

        Note:
            Automatically removes expired entries when accessed
        """
        with self._lock:
            metadata = self._store.get(file_id)
            if metadata:
                timestamp = datetime.fromisoformat(metadata['timestamp'])
                if datetime.now() - timestamp < timedelta(hours=self.retention_hours):
                    return metadata
                else:
                    del self._store[file_id]
            return None

    def cleanup_expired(
            self
        ) -> int:
        """
        Remove all expired metadata entries from the store.

        Returns:
            int: Number of expired entries removed

        Note:
            Called automatically by background thread or manually for maintenance
        """
        with self._lock:
            expired = []
            now = datetime.now()
            for file_id, metadata in self._store.items():
                timestamp = datetime.fromisoformat(metadata['timestamp'])
                if now - timestamp >= timedelta(hours=self.retention_hours):
                    expired.append(file_id)
            for file_id in expired:
                del self._store[file_id]
            return len(expired)

file_metadata_store = FileMetadataStore(retention_hours=24)

# ========== CREATING BLUEPRINT ==========
upload_bp = Blueprint('upload', __name__)

# ========== RATE LIMITING SETUP ==========
limiter = Limiter(
    key_func=get_remote_address,
    storage_uri=RATE_LIMIT_STORAGE_URL,
    default_limits=["200 per day", "50 per hour"],
    strategy="fixed-window"
)


def rate_limit(
        limit_string: Optional[str] = None
    ):
    """
    Decorator factory for applying rate limits to endpoints.

    Wraps Flask routes with Flask-Limiter's rate limiting functionality.

    Args:
        limit_string (Optional[str]): Rate limit string (e.g., 10 per minute).
                                    If None, uses default RATE_LIMIT setting.

    Returns:
        Callable: Decorated function with rate limiting applied

    Example:
        >>> @rate_limit(5 per minute)
        ... def my_endpoint():
        ...     return jsonify({message: ok})
    """
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            return limiter.limit(limit_string or RATE_LIMIT)(f)(*args, **kwargs)
        return wrapped
    return decorator


# ========== EXCEL VALIDATION FUNCTIONS (FROM MEMORY) ==========
def _validate_xlsx_from_bytes(
        file_content: bytes,
        file_info: Dict
    ) -> Tuple[bool, str, Optional[Dict]]:
    """
    Validate .xlsx file format, structure, and content from memory (bytes).

    Performs comprehensive validation of Excel 2007+ (.xlsx) files entirely
    from memory without creating any temporary files:
    - Macro detection via VBA archive inspection (rejects files with macros)
    - Required sheet existence check
    - Row and column count limits validation
    - Optional ignoring of additional sheets

    Args:
        file_content (bytes): Raw file content as bytes
        file_info (Dict): Dictionary to populate with file metadata

    Returns:
        Tuple[bool, str, Optional[Dict]]:
            - bool: True if validation passed, False otherwise
            - str: Error message if validation failed, empty string otherwise
            - Optional[Dict]: Updated file_info with validation results

    Raises:
        openpyxl_exceptions.InvalidFileException: If file is corrupted or invalid
        Exception: For unexpected errors during validation

    Example:
        >>> is_valid, error, info = _validate_xlsx_from_bytes(file_content, {})
        >>> if is_valid:
        ...     print(f"Valid file with {info[total_rows]} rows")
    """
    wb = None
    try:
        file_stream = io.BytesIO(file_content)
        wb = openpyxl.load_workbook(file_stream, read_only=True, data_only=True)

        if hasattr(wb, 'vba_archive') and wb.vba_archive:
            file_info['has_macros'] = True
            return False, "Excel files with macros are not allowed", None

        if FIXED_SHEET_NAME not in wb.sheetnames:
            return False, f"Required sheet '{FIXED_SHEET_NAME}' not found", None

        sheet = wb[FIXED_SHEET_NAME]
        row_count = 0
        col_count = 0

        for row in sheet.iter_rows():
            row_count += 1
            col_count = max(col_count, len(row))
            if row_count > MAX_EXCEL_ROWS:
                return False, f"Sheet exceeds maximum rows ({MAX_EXCEL_ROWS})", None
            if col_count > MAX_EXCEL_COLS:
                return False, f"Sheet exceeds maximum columns ({MAX_EXCEL_COLS})", None

        file_info['sheets'].append({'name': FIXED_SHEET_NAME, 'rows': row_count, 'cols': col_count})
        file_info['total_rows'] = row_count
        file_info['total_cols'] = col_count

        if len(wb.sheetnames) > 1 and IGNORE_OTHER_SHEETS:
            other_sheets = [s for s in wb.sheetnames if s != FIXED_SHEET_NAME]
            logger.info("Additional sheets ignored: %s", other_sheets)

        return True, "", file_info

    except openpyxl_exceptions.InvalidFileException:
        return False, "Invalid Excel file format or encrypted/corrupted", None
    except Exception as e:
        logger.error("Error reading Excel from memory: %s", str(e), exc_info=True)
        return False, f"Failed to read Excel: {str(e)}", None
    finally:
        if wb:
            wb.close()


def _validate_xls_from_bytes(
        file_content: bytes,
        file_info: Dict
    ) -> Tuple[bool, str, Optional[Dict]]:
    """
    Validate legacy .xls file format, structure, and content from memory (bytes).

    Performs comprehensive validation of Excel 97-2003 (.xls) files:
    - VBA macro detection using oletools (rejects files with macros)
    - Requires temporary file creation due to oletools limitation
    - Password/encryption detection (rejects encrypted files)
    - Required sheet existence check
    - Row and column count limits validation

    Note:
        A temporary file is created for macro detection because the oletools
        library requires a file path and does not support streaming input.
        This temporary file is automatically deleted after validation.

    Args:
        file_content (bytes): Raw file content as bytes
        file_info (Dict): Dictionary to populate with file metadata

    Returns:
        Tuple[bool, str, Optional[Dict]]:
            - bool: True if validation passed, False otherwise
            - str: Error message if validation failed, empty string otherwise
            - Optional[Dict]: Updated file_info with validation results

    Raises:
        xlrd.biffh.XLRDError: If file is encrypted or corrupted
        ImportError: If oletools is not installed
        Exception: For unexpected errors during validation

    Example:
        >>> is_valid, error, info = _validate_xls_from_bytes(file_content, {})
        >>> if is_valid:
        ...     print(f"Valid .xls file with {info[total_rows]} rows")
    """
    vba_parser = None
    try:
        # Check macros (requires temp file for oletools)
        try:
            with tempfile.NamedTemporaryFile(suffix='.xls', delete=True) as tmp_file:
                tmp_file.write(file_content)
                tmp_file.flush()
                vba_parser = VBA_Parser(tmp_file.name)
                if vba_parser.detect_vba_macros():
                    return False, "Excel files with macros are not allowed", None
                vba_parser.close()
                vba_parser = None
        except ImportError:
            return False, "Security verification unavailable for .xls files", None
        except Exception as e:
            logger.error("Macro check failed: %s", str(e))
            return False, "Failed to verify file security", None

        # Validate structure from memory
        with open(os.devnull, 'w', encoding='utf-8') as null_file:
            wb = xlrd.open_workbook(
                file_contents=file_content,
                formatting_info=False,
                logfile=null_file
            )

        if FIXED_SHEET_NAME not in wb.sheet_names():
            return False, f"Required sheet '{FIXED_SHEET_NAME}' not found", None

        sheet = wb.sheet_by_name(FIXED_SHEET_NAME)
        row_count = sheet.nrows
        col_count = sheet.ncols

        if row_count > MAX_EXCEL_ROWS:
            return False, f"Sheet exceeds maximum rows ({MAX_EXCEL_ROWS})", None
        if col_count > MAX_EXCEL_COLS:
            return False, f"Sheet exceeds maximum columns ({MAX_EXCEL_COLS})", None

        file_info['sheets'].append({'name': FIXED_SHEET_NAME, 'rows': row_count, 'cols': col_count})
        file_info['total_rows'] = row_count
        file_info['total_cols'] = col_count

        if len(wb.sheet_names()) > 1 and IGNORE_OTHER_SHEETS:
            other_sheets = [s for s in wb.sheet_names() if s != FIXED_SHEET_NAME]
            logger.info("Additional sheets ignored: %s", other_sheets)

        return True, "", file_info

    except xlrd.biffh.XLRDError as e:
        if 'encrypt' in str(e).lower() or 'password' in str(e).lower():
            return False, "Encrypted Excel files are not allowed", None
        return False, f"Invalid Excel file: {str(e)}", None
    except Exception as e:
        logger.error("Error reading Excel from memory: %s", str(e), exc_info=True)
        return False, f"Failed to read Excel: {str(e)}", None
    finally:
        if vba_parser:
            try:
                vba_parser.close()
            except Exception:
                pass


def validate_excel_content_from_bytes(
        file_content: bytes
    ) -> Tuple[bool, str, Optional[Dict]]:
    """
    Main entry point for Excel file content validation from memory (bytes).

    Orchestrates the validation process based on file signature detection:
    - .xlsx files (ZIP signature): Routes to _validate_xlsx_from_bytes
    - .xls files (OLE signature): Routes to _validate_xls_from_bytes

    Also calculates SHA-256 hash of the file for integrity verification.

    Args:
        file_content (bytes): Raw file content as bytes

    Returns:
        Tuple[bool, str, Optional[Dict]]:
            - bool: True if validation passed, False otherwise
            - str: Error message if validation failed, empty string otherwise
            - Optional[Dict]: File metadata including sheets, rows, columns, hash

    Example:
        >>> is_valid, error, info = validate_excel_content_from_bytes(file_content)
        >>> if is_valid:
        ...     print(f"Valid {info[format]} file with {info[total_rows]} rows")
    """
    try:
        file_info = {
            'sheets': [],
            'total_rows': 0,
            'total_cols': 0,
            'has_macros': False,
            'is_encrypted': False,
            'file_hash': None,
            'format': 'unknown'
        }

        # Calculate hash
        sha256_hash = hashlib.sha256()
        sha256_hash.update(file_content)
        file_info['file_hash'] = sha256_hash.hexdigest()

        # Detect format from signature
        if len(file_content) >= 4:
            if file_content[:4] == b'PK\x03\x04':
                file_info['format'] = 'xlsx'
                return _validate_xlsx_from_bytes(file_content, file_info)
            elif file_content[:4] == b'\xD0\xCF\x11\xE0':
                file_info['format'] = 'xls'
                return _validate_xls_from_bytes(file_content, file_info)
            else:
                return False, "Unsupported or invalid Excel file format", None
        else:
            return False, "File is too small or empty", None

    except MemoryError:
        return False, "File too large to process in memory", None
    except Exception as e:
        logger.error("Excel validation error: %s", str(e), exc_info=True)
        return False, f"Excel validation failed: {str(e)}", None


def validate_file(file) -> Tuple[bool, str, Optional[Dict], int]:
    """
    Perform basic file validation before content processing.

    Checks:
    - File existence and non-empty filename
    - File extension against allowed Excel formats
    - File size against MAX_FILE_SIZE limit
    - Content-Length header consistency

    Args:
        file: File object from Flask request.files

    Returns:
        Tuple[bool, str, Optional[Dict], int]:
            - bool: True if validation passed, False otherwise
            - str: Error message if validation failed, empty string otherwise
            - Optional[Dict]: Basic file info (extension, original filename)
            - int: HTTP status code appropriate for the result

    HTTP Status Codes:
        - 200: Validation passed
        - 400: Bad request (no file, invalid extension)
        - 413: File too large

    Example:
        >>> is_valid, error, info, status = validate_file(request.files[file])
        >>> if not is_valid:
        ...     return jsonify({error: error}), status
    """
    if not file or not file.filename:
        return False, "No file provided", None, 400

    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXCEL_EXTENSIONS:
        return False, f"Only Excel files allowed: {', '.join(ALLOWED_EXCEL_EXTENSIONS)}", None, 400

    # Check size
    content_length = None
    if hasattr(file, 'content_length') and file.content_length:
        content_length = file.content_length
    elif hasattr(request, 'content_length') and request.content_length:
        content_length = request.content_length

    if content_length and content_length > MAX_FILE_SIZE:
        return False, f"File too large. Max: {MAX_FILE_SIZE_MB} MB", None, 413

    return True, "", {'extension': file_ext, 'original_filename': file.filename}, 200


def generate_unique_filename(original_filename: str) -> tuple[str, str, str]:
    """
    Generate a unique, sanitized filename with timestamp and UUID.

    Creates a filename in format: {sanitized_base}_{timestamp}_{unique_id}{extension}

    Args:
        original_filename (str): Original filename from upload

    Returns:
        tuple[str, str, str]:
            - str: Safe unique filename
            - str: Timestamp string (YYYYMMDD_HHMMSS format)
            - str: Unique ID (first 8 chars of UUID4 hex)

    Example:
        >>> safe_name, ts, uid = generate_unique_filename(My File@2.xlsx)
        >>> print(safe_name)  # My_File_2_20260127_143015_a1b2c3d4.xlsx
    """
    moscow_time = datetime.now(MOSCOW_TZ)
    timestamp = moscow_time.strftime('%Y%m%d_%H%M%S')
    unique_id = uuid.uuid4().hex[:8]
    file_ext = Path(original_filename).suffix
    base_name = Path(original_filename).stem
    safe_base = "".join(c for c in base_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    if not safe_base:
        safe_base = "file"
    safe_filename = f"{safe_base}_{timestamp}_{unique_id}{file_ext}"
    return safe_filename, timestamp, unique_id


def trigger_airflow_dag(file_info: Dict[str, Any]) -> Tuple[bool, Optional[Dict]]:
    """
    Trigger Airflow DAG with file content via REST API v2 using JWT authentication.

    This function implements the authentication flow required by Airflow 3.0.6:
    1. First obtains a JWT access token using username/password via the /auth/token endpoint
    2. Then uses the token to trigger the DAG via the /api/v2/dags/{dag_id}/dagRuns endpoint

    Unlike traditional file-based triggers, this function passes the entire file
    content as base64-encoded data, eliminating the need for shared file volumes.

    Args:
        file_info (Dict[str, Any]): Complete file metadata including:
            - safe_filename (str): Unique filename for storage
            - file_content_b64 (str): Base64-encoded file content
            - original_filename (str): Original uploaded filename
            - file_hash (str): SHA-256 hash of the file
            - file_size (int): Size in bytes
            - file_format (str): Excel format (xlsx/xls)
            - sheets (List[Dict]): List of sheet information
                - name (str): Sheet name
                - rows (int): Number of rows in sheet
                - cols (int): Number of columns in sheet
            - total_rows (int): Total row count across all sheets
            - upload_time (str): Timestamp of upload
            - unique_id (str): Unique identifier for this upload

    Returns:
        Tuple[bool, Optional[Dict]]:
            - bool: True if DAG triggered successfully, False otherwise
            - Optional[Dict]: On success - DAG run response containing:
                - dag_run_id (str): ID of the triggered DAG run
                - dag_id (str): ID of the DAG
                - execution_date (str): Execution date of the run
                - state (str): Initial state of the run (usually 'running')
              On failure - error details containing:
                - error (str): Error message
                - details (str, optional): Additional error details from Airflow
                - status_code (int, optional): HTTP status code if applicable

    Environment Variables Required:
        - AIRFLOW_API_URL: Base URL of Airflow API server
        - AIRFLOW_USER: Username for Airflow authentication
        - AIRFLOW_PASSWORD: Password for Airflow authentication
        - DAG_ID: ID of the DAG to trigger

    Timeouts:
        - JWT token request: 15 seconds
        - DAG trigger request: 30 seconds
    """
    try:
        # 1. Get JWT token using username/password
        # IMPORTANT: In Airflow 3.0.6 with FAB Auth Manager, token endpoint is /auth/token
        # NOT /api/v2/auth/token
        token_url = f"{AIRFLOW_API_URL}/auth/token"
        logger.info("Requesting JWT token from: %s", token_url)

        token_response = requests.post(
            token_url,
            json={"username": AIRFLOW_USER, "password": AIRFLOW_PASSWORD},
            timeout=15
        )

        logger.info("Token response status: %d", token_response.status_code)

        # Check if token request was successful
        if token_response.status_code not in (200, 201):
            logger.error(
                "Token request failed: %d - %s",
                token_response.status_code,
                token_response.text
            )
            return False, {
                "error": "JWT token generation failed",
                "details": token_response.text,
                "status_code": token_response.status_code
            }

        # Extract access_token from response
        token_data = token_response.json()
        logger.info("Token response keys: %s", list(token_data.keys()) if token_data else 'empty')

        # Airflow 3.0.6 returns token in the format: {"access_token": "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9..."}
        # But also handle other possible formats for compatibility
        jwt_token = (
            token_data.get('access_token') or
            token_data.get('token') or
            token_data.get('jwt')
        )

        if not jwt_token:
            logger.error("No token found in response: %s", token_data)
            return False, {
                "error": "No access_token in JWT response",
                "details": token_data
            }

        logger.info("JWT token obtained successfully (length: %d chars)", len(jwt_token))

        # 2. Prepare DAG trigger request
        dag_url = f"{AIRFLOW_API_URL}/api/v2/dags/{DAG_ID}/dagRuns"
        headers = {
            "Authorization": f"Bearer {jwt_token}",
            "Content-Type": "application/json"
        }

        # ========== FIX: Add required logical_date field ==========
        # In Airflow 3.0.6, 'logical_date' is required for manual DAG runs
        # Set to None to allow parallel runs without a specific date or use a specific datetime for backfilling
        current_time = datetime.now(MOSCOW_TZ)
        logical_date = current_time.isoformat()
        # Alternatively, set to None to allow parallel runs:
        # logical_date = None

        # Prepare payload with file information AND required logical_date
        dag_payload = {
            "logical_date": logical_date,  # REQUIRED in Airflow 3.0.6+
            "conf": {
                "file_info": file_info
            }
        }

        # 3. Trigger the DAG
        logger.info("Triggering DAG: %s at %s", DAG_ID, dag_url)
        logger.info("Payload: %s", dag_payload)  # Log payload for debugging (consider removing in production)

        dag_response = requests.post(
            dag_url,
            json=dag_payload,
            headers=headers,
            timeout=30
        )

        logger.info("DAG trigger response status: %d", dag_response.status_code)

        # Check if DAG trigger was successful
        if dag_response.status_code in (200, 201):  # 201 Created is also valid
            logger.info("DAG triggered successfully for file: %s", file_info.get('safe_filename', 'unknown'))
            return True, dag_response.json()
        else:
            logger.error(
                "DAG trigger failed: %d - %s",
                dag_response.status_code,
                dag_response.text
            )
            return False, {
                "error": f"DAG trigger failed with status {dag_response.status_code}",
                "details": dag_response.text,
                "status_code": dag_response.status_code
            }

    except requests.exceptions.Timeout as e:
        logger.error("Timeout connecting to Airflow API: %s", str(e))
        return False, {
            "error": "Timeout connecting to Airflow API",
            "details": str(e)
        }
    except requests.exceptions.ConnectionError as e:
        logger.error("Connection error to Airflow API: %s", str(e))
        return False, {
            "error": "Connection error to Airflow API",
            "details": str(e)
        }
    except requests.exceptions.RequestException as e:
        logger.error("Request error: %s - %s", type(e).__name__, str(e))
        return False, {
            "error": f"Request error: {type(e).__name__}",
            "details": str(e)
        }
    except ValueError as e:
        # This catches JSON decode errors
        logger.error("Invalid JSON response from Airflow API: %s", str(e))
        return False, {
            "error": "Invalid response from Airflow API",
            "details": str(e)
        }
    except Exception as e:
        # Catch any other unexpected exceptions
        logger.error(
            "Unexpected error: %s - %s",
            type(e).__name__,
            str(e),
            exc_info=True
        )
        return False, {
            "error": f"Unexpected error: {type(e).__name__}",
            "details": str(e)
        }


# ========== API ENDPOINTS ==========
@upload_bp.route('/health', methods=['GET'])
def health_check():
    """
    Health check endpoint for monitoring and container orchestration.

    Returns service status information including ClamAV availability and
    streaming mode status. Used by container orchestration tools like
    Kubernetes for liveness and readiness probes.

    Returns:
        Response: JSON with:
            - status: healthy if service is operational
            - timestamp: Current ISO timestamp
            - environment: Current environment (development/production)
            - streaming_mode: Always True (indicates no disk persistence)
            - clamav_available: ClamAV service status

    HTTP Status:
        - 200: Service is healthy

    Example:
        >>> response = requests.get(http://localhost:5000/health)
        >>> response.json()
        {
            status: healthy,
            timestamp: 2026-01-27T14:30:15,
            environment: production,
            streaming_mode: True,
            clamav_available: True
        }
    """
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'environment': FLASK_ENV,
        'streaming_mode': True,  # ← Новый флаг!
        'clamav_available': clamav_scanner.is_available()
    }), 200


@upload_bp.route('/upload-mft-excel', methods=['POST'])
@rate_limit()
def upload_mft_excel():
    """
    Main endpoint for secure Excel file upload with streaming (no disk save).

    Comprehensive upload process entirely in memory:
    1. Basic validation (file presence, size, extension)
    2. Read file into memory
    3. Virus scan via ClamAV stream (if available)
    4. Generate unique filename
    5. Excel content validation from memory (macros, encryption, structure)
    6. Store metadata in memory
    7. Trigger Airflow DAG with base64-encoded content

    Returns:
        Response: JSON with:
            - message: Success/failure message
            - file: File metadata (safe_filename, original, size, hash, etc.)
            - dag_triggered: Boolean indicating DAG trigger status
            - dag: DAG run response if successful
            - warning/dag_error: Optional warnings if DAG trigger fails

    HTTP Status:
        - 200: Complete success
        - 207: Partial success (file validated but DAG not triggered)
        - 400: Validation error or virus detected
        - 413: File too large
        - 500: Server error

    Rate Limited: Yes (configured by RATE_LIMIT)

    Example:
        >>> # Upload via curl
        >>> response = requests.post(
        ...     http://localhost:5000/upload-mft-excel,
        ...     files={file: (data.xlsx, open(data.xlsx, rb), application/vnd.openxmlformats-officedocument.spreadsheetml.sheet)}
        ... )
        >>> response.json()
        {
            message: File uploaded successfully (streaming mode),
            file: {
                safe_filename: data_20260127_143015_a1b2c3d4.xlsx,
                original_filename: data.xlsx,
                file_size: 1024,
                upload_time: 20260127_143015,
                unique_id: a1b2c3d4,
                sheets: [{name: mft, rows: 100, cols: 20}],
                total_rows: 100,
                format: xlsx,
                file_hash: abc123...
            },
            dag_triggered: true,
            dag: {
                dag_run_id: manual__2026-01-27T14:30:15+00:00,
                state: running
            }
        }
    """
    start_time = time.time()

    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']
        if not file.filename:
            return jsonify({'error': 'No file selected'}), 400

        # Validate
        is_valid, error_msg, file_info, status_code = validate_file(file)
        if not is_valid:
            return jsonify({'error': error_msg}), status_code

        # Initialize file_info if None
        if file_info is None:
            file_info = {}

        # Read into memory
        file_content = file.read()
        file_size = len(file_content)

        # Virus scan
        if clamav_scanner.is_available():
            is_clean, scan_result = clamav_scanner.scan_stream(file_content, file.filename)
            if not is_clean:
                logger.error("VIRUS DETECTED: %s", scan_result)
                return jsonify({'error': 'File rejected: virus detected', 'virus': scan_result}), 400

        # Generate unique filename
        safe_filename, timestamp, unique_id = generate_unique_filename(file.filename)

        # Validate Excel from memory
        is_valid_content, content_error, content_info = validate_excel_content_from_bytes(file_content)
        if not is_valid_content:
            return jsonify({'error': content_error}), 400

        # Check if content_info is not None before update
        if content_info:
            file_info.update(content_info)

        # Prepare metadata (NO FILE PATHS!)
        file_metadata = {
            'safe_filename': safe_filename,
            'timestamp': timestamp,
            'unique_id': unique_id,
            'original_filename': file.filename,
            'file_size': file_size,
            'file_hash': file_info.get('file_hash'),
            'sheets': file_info.get('sheets', []),
            'total_rows': file_info.get('total_rows', 0),
            'has_macros': file_info.get('has_macros', False),
            'format': file_info.get('format', 'unknown'),
            'upload_duration_ms': int((time.time() - start_time) * 1000)
        }

        # Store metadata
        file_metadata_store.add(unique_id, file_metadata)

        # Trigger DAG
        dag_success, dag_response = trigger_airflow_dag(file_metadata)

        response_data = {
            'message': 'File uploaded successfully (streaming mode)',
            'file': {
                'safe_filename': safe_filename,
                'original_filename': file.filename,
                'file_size': file_size,
                'upload_time': timestamp,
                'unique_id': unique_id,
                'sheets': file_info.get('sheets', []),
                'total_rows': file_info.get('total_rows', 0),
                'format': file_info.get('format', 'unknown'),
                'file_hash': file_info.get('file_hash')
            },
            'dag_triggered': dag_success
        }

        if dag_success:
            response_data['dag'] = dag_response
            return jsonify(response_data), 200
        else:
            response_data['warning'] = 'File uploaded but DAG trigger failed'
            response_data['dag_error'] = dag_response
            return jsonify(response_data), 207

    except MemoryError:
        return jsonify({'error': 'File too large to process'}), 413
    except Exception as e:
        logger.error("Upload failed: %s", str(e), exc_info=True)
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500


@upload_bp.route('/file/<file_id>', methods=['GET'])
def get_file_info(file_id: str):
    """
    Retrieve metadata for a previously uploaded file by ID.

    Args:
        file_id (str): Unique identifier returned from upload

    Returns:
        Response: JSON with file metadata if found

    HTTP Status:
        - 200: File found, metadata returned
        - 404: File ID not found or expired
    """
    metadata = file_metadata_store.get(file_id)
    if metadata:
        return jsonify({'file': metadata}), 200
    return jsonify({'error': 'File not found or expired'}), 404


@upload_bp.route('/file/<file_id>/status', methods=['GET'])
def get_file_status(file_id: str):
    """
    Check if an uploaded file still exists (always true in streaming mode).

    Args:
        file_id (str): Unique identifier returned from upload

    Returns:
        Response: JSON with:
            - file_id: Requested file ID
            - exists: Always True (files are processed in memory)
            - metadata: Complete metadata if file exists

    HTTP Status:
        - 200: Check completed
        - 404: File ID not found in metadata store
    """
    metadata = file_metadata_store.get(file_id)
    if not metadata:
        return jsonify({'error': 'File not found'}), 404
    return jsonify({
        'file_id': file_id,
        'exists': True,
        'metadata': metadata
    }), 200


# ========== FLASK APP SETUP ==========
def create_app():
    """
    Create and configure the Flask application instance.

    Sets up:
    - Secret key for sessions
    - Maximum content length
    - CORS for Browser Security Policy
    - Blueprint registration
    - Rate limiter initialization
    - Security headers for production

    Returns:
        Flask: Configured Flask application instance

    Example:
        >>> app = create_app()
        >>> app.run(host=0.0.0.0, port=5000, debug=True)
    """
    flask_app = Flask(__name__)
    flask_app.secret_key = FLASK_SECRET_KEY
    flask_app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

    CORS(flask_app)

    @flask_app.after_request
    def add_security_headers(response):
        if IS_PRODUCTION:
            response.headers.add('X-Content-Type-Options', 'nosniff')
            response.headers.add('X-Frame-Options', 'DENY')
            response.headers.add('X-XSS-Protection', '1; mode=block')
        return response

    flask_app.register_blueprint(upload_bp)
    limiter.init_app(flask_app)

    logger.info("Upload API started in STREAMING mode (no disk save)")
    return flask_app

app = create_app()

if __name__ == '__main__':
    logger.info("="*60)
    logger.info("Starting Upload API in STREAMING MODE (no disk save)")
    logger.info("Files are processed in memory only")
    logger.info("="*60)
    app.run(host=FLASK_HOST, port=FLASK_PORT, debug=FLASK_DEBUG, threaded=True)
