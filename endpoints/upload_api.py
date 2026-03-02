# pylint: disable=too-many-lines
# pylint: disable=wrong-import-position
"""
Upload API module for secure Excel file uploads to Airflow shared volume with DAG triggering.

This module provides a Flask-based REST API for uploading Excel files with comprehensive
security features including virus scanning via ClamAV, content validation, rate limiting,
and automatic DAG triggering in Apache Airflow. It implements production-ready file
handling with backup creation, quarantine for infected files, and background cleanup.

Key Features:
    - Secure file upload with virus scanning via ClamAV integration
    - Excel content validation (.xlsx and .xls) with macro detection
    - Rate limiting to prevent DDoS attacks
    - Thread-safe metadata storage with auto-cleanup
    - Automatic DAG triggering after successful upload
    - Health check and statistics endpoints
    - Background cleanup of expired files
    - Comprehensive error handling and logging

Environment Variables:
    FLASK_SECRET_KEY: Secret key for Flask sessions (required)
    FLASK_HOST: Host to bind the server (default: 0.0.0.0)
    FLASK_PORT: Port to bind the server (default: 5000)
    FLASK_ENV: Environment (development/production)
    MAX_FILE_SIZE_MB: Maximum allowed file size in MB (default: 5)
    FILE_RETENTION_DAYS: Days to retain uploaded files (default: 1)
    RATE_LIMIT: Rate limit string (default: '10 per minute')
    AIRFLOW_API_URL: Airflow API endpoint URL
    DAG_ID: DAG ID to trigger after upload
    FIXED_SHEET_NAME: Required sheet name in Excel files (default: 'mft')

Endpoints:
    GET /health - Health check for monitoring
    GET /stats - Upload statistics (admin only)
    POST /cleanup - Manual cleanup trigger (admin only)
    POST /upload-mft-excel - Main file upload endpoint
    GET /file/<file_id> - Get file metadata by ID
    GET /file/<file_id>/status - Check file existence

Version: 1.0.0
Compatibility: Python 3.12.3
Maintainer: PLD Engineering Center
Created: 2026-02-26
Last Modified: 2026-02-26
License: MIT
Status: Production
"""
# Standard library imports
import os
import sys
import tempfile
import shutil
import uuid
import hashlib
import time
import threading
from pathlib import Path
from datetime import datetime, timedelta
from functools import wraps
from threading import Lock
from typing import Tuple, Optional, Dict, Any

# Third-party imports
import requests
import xlrd
import openpyxl
from oletools.olevba import VBA_Parser
from flask import Blueprint, Flask, jsonify, request, session
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from dotenv import load_dotenv
from openpyxl.utils import exceptions as openpyxl_exceptions

# The relative path to the root project directory
try:
    PROJECT_ROOT = Path(__file__).resolve().parents[1]
except NameError:
    # If __file__ is not defined (in exec() or interactive mode)
    PROJECT_ROOT = Path("/opt/airflow")

# Add the project path to sys.path
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

# Local imports
from config import get_logger
from config.clamav_service import clamav_scanner

# Logger setup
logger = get_logger(__name__)

# Load environment variables
env_path = PROJECT_ROOT / '.env'
load_dotenv(dotenv_path=env_path)

# ========== CONFIGURATION ==========

# Flask/Upload API configuration
FLASK_SECRET_KEY = os.getenv('FLASK_SECRET_KEY')
if not FLASK_SECRET_KEY:
    raise RuntimeError("FLASK_SECRET_KEY must be set in .env file")

FLASK_HOST = os.getenv('FLASK_HOST', '0.0.0.0')
FLASK_PORT = int(os.getenv('FLASK_PORT', '5000'))
FLASK_DEBUG = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'

# Defining the environment (default is development)
FLASK_ENV = os.getenv('FLASK_ENV', 'development')
IS_PRODUCTION = FLASK_ENV == 'production'

# Airflow configuration
AIRFLOW_API_URL = os.getenv('AIRFLOW_API_URL', 'http://airflow-webserver:8080/api/v1')
AIRFLOW_USER = os.getenv('AIRFLOW_USER', 'admin')
AIRFLOW_PASSWORD = os.getenv('AIRFLOW_PASSWORD', 'airflow')
DAG_ID = os.getenv('DAG_ID', 'mft_etl_pipeline')

# File upload settings
MAX_FILE_SIZE_MB = int(os.getenv('MAX_FILE_SIZE_MB', '5'))
MAX_FILE_SIZE = MAX_FILE_SIZE_MB * 1024 * 1024
FILE_RETENTION_DAYS = int(os.getenv('FILE_RETENTION_DAYS', '1'))

# Rate limiting settings
RATE_LIMIT = os.getenv('RATE_LIMIT', '10 per minute')
RATE_LIMIT_STORAGE_URL = os.getenv('RATE_LIMIT_STORAGE_URL', 'memory://')

# Excel validation settings
MAX_EXCEL_ROWS = int(os.getenv('MAX_EXCEL_ROWS', '1000000'))
MAX_EXCEL_COLS = int(os.getenv('MAX_EXCEL_COLS', '1000'))
extensions_str = os.getenv('ALLOWED_EXCEL_EXTENSIONS', '.xlsx,.xls')
ALLOWED_EXCEL_EXTENSIONS = set(extensions_str.split(','))
FIXED_SHEET_NAME = os.getenv('FIXED_SHEET_NAME', 'mft')
IGNORE_OTHER_SHEETS = os.getenv('IGNORE_OTHER_SHEETS', 'true').lower() == 'true'

# Paths
AIRFLOW_PROJ_DIR = os.getenv('AIRFLOW_PROJ_DIR', '.')
AIRFLOW_DATA_DIR = Path(os.getenv('AIRFLOW_DATA_DIR', '/opt/airflow/data'))
HOST_DATA_DIR = Path(AIRFLOW_PROJ_DIR) / 'data'

# Subdirectories
UPLOAD_SUBDIR = 'uploads'
BACKUP_SUBDIR = 'backups'
PROCESSED_SUBDIR = 'processed'
QUARANTINE_SUBDIR = 'quarantine'

# Full paths
UPLOAD_DIR = AIRFLOW_DATA_DIR / UPLOAD_SUBDIR
BACKUP_DIR = AIRFLOW_DATA_DIR / BACKUP_SUBDIR
PROCESSED_DIR = AIRFLOW_DATA_DIR / PROCESSED_SUBDIR
QUARANTINE_DIR = AIRFLOW_DATA_DIR / QUARANTINE_SUBDIR

# Host paths (for creating directories)
HOST_UPLOAD_DIR = UPLOAD_DIR
HOST_BACKUP_DIR = BACKUP_DIR
HOST_PROCESSED_DIR = PROCESSED_DIR
HOST_QUARANTINE_DIR = QUARANTINE_DIR

# ========== THREAD-SAFE STORAGE FOR FILE METADATA ==========
class FileMetadataStore:
    """
    Thread-safe storage for file metadata with automatic expiration and cleanup.
    
    This class provides a thread-safe dictionary-like store for file metadata
    with timestamp-based expiration. It automatically removes entries older
    than the specified retention period.
    
    Attributes:
        retention_hours (int): Number of hours to retain metadata entries
        _store (Dict[str, Dict[str, Any]]): Internal metadata storage
        _lock (Lock): Thread lock for safe concurrent access
    
    Methods:
        add: Add file metadata with timestamp
        get: Retrieve metadata if not expired
        cleanup_expired: Remove expired entries
    
    Example:
        >>> store = FileMetadataStore(retention_hours=24)
        >>> store.add("file123", {"name": "data.xlsx", "size": 1024})
        >>> metadata = store.get("file123")
    """
    def __init__(self, retention_hours: int = 24):
        """
        Initialize the metadata store with retention period.
    
        Args:
            retention_hours (int): Number of hours to retain metadata entries.
                                   Defaults to 24 hours.
        """
        self._store: Dict[str, Dict[str, Any]] = {}
        self._lock = Lock()
        self.retention_hours = retention_hours

    def add(self, file_id: str, metadata: Dict[str, Any]) -> None:
        """
        Add file metadata to the store with current timestamp.
        
        Args:
            file_id (str): Unique identifier for the file
            metadata (Dict[str, Any]): File metadata dictionary to store
            
        Note:
            Automatically adds 'timestamp' and 'file_id' keys to metadata
        """
        with self._lock:
            metadata['timestamp'] = datetime.now().isoformat()
            metadata['file_id'] = file_id
            self._store[file_id] = metadata

    def get(self, file_id: str) -> Optional[Dict[str, Any]]:
        """
        Retrieve file metadata if it exists and hasn't expired.
        
        Args:
            file_id (str): Unique identifier for the file
            
        Returns:
            Optional[Dict[str, Any]]: Metadata dictionary if found and not expired,
                                    None otherwise
            
        Note:
            Automatically removes expired entries when accessed
        """
        with self._lock:
            metadata = self._store.get(file_id)
            if metadata:
                timestamp = datetime.fromisoformat(metadata['timestamp'])
                if datetime.now() - timestamp < timedelta(hours=self.retention_hours):
                    return metadata
                else:
                    del self._store[file_id]
            return None

    def cleanup_expired(self) -> int:
        """
        Remove all expired metadata entries from the store.
        
        Returns:
            int: Number of expired entries removed
            
        Note:
            Called automatically by background cleanup thread
        """
        with self._lock:
            expired = []
            now = datetime.now()
            for file_id, metadata in self._store.items():
                timestamp = datetime.fromisoformat(metadata['timestamp'])
                if now - timestamp >= timedelta(hours=self.retention_hours):
                    expired.append(file_id)

            for file_id in expired:
                del self._store[file_id]

            return len(expired)

# Initialize metadata store
file_metadata_store = FileMetadataStore(retention_hours=FILE_RETENTION_DAYS * 24)

# ========== CREATING DIRECTORIES ==========
for directory in [HOST_UPLOAD_DIR, HOST_BACKUP_DIR, HOST_PROCESSED_DIR, HOST_QUARANTINE_DIR]:
    try:
        directory.mkdir(parents=True, exist_ok=True)
        logger.info("The directory was created successfully: %s", directory)

        # Checking the recording rights
        test_file = directory / '.write_test'
        test_file.touch()
        test_file.unlink()
        logger.info("Directory is writable: %s", directory)

    except PermissionError as e:
        logger.error("No access rights for %s: %s", directory, e)
        raise

    except OSError as e:
        logger.error("File system error for %s: %s", directory, e)
        raise

    except Exception as unexpected_error:
        logger.error("Unexpected error for %s: %s", directory, unexpected_error)
        raise

# ========== CLAMAV CHECK ==========
if not clamav_scanner.is_available():
    logger.warning("ClamAV is not available! Files will be uploaded WITHOUT virus scanning!")
    logger.warning("Check if clamav container is running")
else:
    logger.info("ClamAV is available and ready")
    clamav_stats = clamav_scanner.get_stats()
    logger.info("Temp directory: %s", clamav_stats.get('temp_dir'))
    if 'temp_stats' in clamav_stats:
        logger.info("Temp files: %s", clamav_stats['temp_stats'].get('files_count', 0))

# ========== RATE LIMITING SETUP ==========
limiter = Limiter(
    key_func=get_remote_address,
    storage_uri=RATE_LIMIT_STORAGE_URL,
    default_limits=["200 per day", "50 per hour"],
    strategy="fixed-window"
)

# ========== CREATING BLUEPRINT ==========
upload_bp = Blueprint('upload', __name__)

# ========== EXCEL VALIDATION FUNCTIONS ==========
def _validate_xlsx(file_path: Path, file_info: Dict) -> Tuple[bool, str, Optional[Dict]]:
    """
    Validate .xlsx file format, structure, and content.
    
    Performs comprehensive validation of Excel 2007+ (.xlsx) files including:
    - Macro detection (rejects files with VBA macros)
    - Required sheet existence check
    - Row and column count limits validation
    - Optional ignoring of additional sheets
    
    Args:
        file_path (Path): Path to the .xlsx file
        file_info (Dict): Dictionary to populate with file metadata
        
    Returns:
        Tuple[bool, str, Optional[Dict]]: 
            - bool: True if validation passed, False otherwise
            - str: Error message if validation failed, empty string otherwise
            - Optional[Dict]: Updated file_info if validation passed, None otherwise
            
    Raises:
        openpyxl_exceptions.InvalidFileException: If file is corrupted or invalid
        Exception: For unexpected errors during validation
    """
    wb = None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

        # Check for macros
        if hasattr(wb, 'vba_archive') and wb.vba_archive:
            file_info['has_macros'] = True
            logger.warning("VBA macros detected in .xlsx file: %s - REJECTED", file_path.name)
            return False, "Excel files with macros are not allowed for security reasons", None

        # Check if required sheet exists
        if FIXED_SHEET_NAME not in wb.sheetnames:
            return False, f"Required sheet '{FIXED_SHEET_NAME}' not found in file", None

        # Get the fixed sheet
        sheet = wb[FIXED_SHEET_NAME]

        # Get dimensions efficiently in read_only mode
        row_count = 0
        col_count = 0

        for row in sheet.iter_rows():
            row_count += 1
            col_count = max(col_count, len(row))

            # Early exit if limits exceeded
            if row_count > MAX_EXCEL_ROWS:
                return (
                    False,
                    f"Sheet exceeds maximum rows ({MAX_EXCEL_ROWS})",
                    None
                )

            if col_count > MAX_EXCEL_COLS:
                return (
                    False,
                    f"Sheet exceeds maximum columns ({MAX_EXCEL_COLS})",
                    None
                )

        file_info['sheets'].append({
            'name': FIXED_SHEET_NAME,
            'rows': row_count,
            'cols': col_count
        })

        file_info['total_rows'] = row_count
        file_info['total_cols'] = col_count

        # Log if there are other sheets (informational only)
        if len(wb.sheetnames) > 1 and IGNORE_OTHER_SHEETS:
            other_sheets = [s for s in wb.sheetnames if s != FIXED_SHEET_NAME]
            logger.info("File contains additional sheets that will be ignored: %s", other_sheets)

        return True, "", file_info

    except openpyxl_exceptions.InvalidFileException:
        logger.warning(
            "Invalid Excel file format (possibly encrypted or corrupted): %s",
            file_path.name
        )
        return False, "Invalid Excel file format or file is encrypted/corrupted", None

    except Exception as unexpected_error:
        logger.error(
            "Unexpected error reading Excel file: %s",
            str(unexpected_error), exc_info=True
        )
        return False, f"Failed to read Excel file: {str(unexpected_error)}", None

    finally:
        if wb:
            wb.close()

def _validate_xls(file_path: Path, file_info: Dict) -> Tuple[bool, str, Optional[Dict]]:
    """
    Validate legacy .xls file format, structure, and content.
    
    Performs comprehensive validation of Excel 97-2003 (.xls) files including:
    - VBA macro detection using oletools (rejects files with macros)
    - Password/encryption detection (rejects encrypted files)
    - Required sheet existence check
    - Row and column count limits validation
    - Optional ignoring of additional sheets
    
    Args:
        file_path (Path): Path to the .xls file
        file_info (Dict): Dictionary to populate with file metadata
        
    Returns:
        Tuple[bool, str, Optional[Dict]]: 
            - bool: True if validation passed, False otherwise
            - str: Error message if validation failed, empty string otherwise
            - Optional[Dict]: Updated file_info if validation passed, None otherwise
            
    Raises:
        xlrd.biffh.XLRDError: If file is encrypted or corrupted
        ImportError: If oletools is not installed
        Exception: For unexpected errors during validation
    """
    wb = None
    vba_parser = None

    try:
        # 1. Check for macros
        try:
            vba_parser = VBA_Parser(str(file_path))

            if vba_parser.detect_vba_macros():
                logger.warning("VBA macros detected in .xls file: %s - REJECTED", file_path.name)
                return False, "Excel files with macros are not allowed for security reasons", None

            vba_parser.close()
            vba_parser = None

        except ImportError:
            logger.error("oletools not installed - cannot verify .xls files for macros")
            return False, "Security verification unavailable for .xls files", None

        except Exception as unexpected_error:
            logger.error("Failed to check macros with oletools: %s", str(unexpected_error))
            return False, "Failed to verify file security", None

        # 2. Validate structure
        with open(os.devnull, 'w', encoding='utf-8') as null_file:
            wb = xlrd.open_workbook(
                filename=str(file_path),
                formatting_info=False,
                logfile=null_file
            )

        # Check if required sheet exists
        if FIXED_SHEET_NAME not in wb.sheet_names():
            return False, f"Required sheet '{FIXED_SHEET_NAME}' not found in file", None

        # Get the fixed sheet
        sheet = wb.sheet_by_name(FIXED_SHEET_NAME)

        row_count = sheet.nrows
        col_count = sheet.ncols

        # Check limits
        if row_count > MAX_EXCEL_ROWS:
            return (
                False,
                f"Sheet exceeds maximum rows ({MAX_EXCEL_ROWS})",
                None
            )

        if col_count > MAX_EXCEL_COLS:
            return (
                False,
                f"Sheet exceeds maximum columns ({MAX_EXCEL_COLS})",
                None
            )

        file_info['sheets'].append({
            'name': FIXED_SHEET_NAME,
            'rows': row_count,
            'cols': col_count
        })

        file_info['total_rows'] = row_count
        file_info['total_cols'] = col_count

        # Log if there are other sheets (informational only)
        if len(wb.sheet_names()) > 1 and IGNORE_OTHER_SHEETS:
            other_sheets = [s for s in wb.sheet_names() if s != FIXED_SHEET_NAME]
            logger.info("File contains additional sheets that will be ignored: %s", other_sheets)

        return True, "", file_info

    except xlrd.biffh.XLRDError as e:
        error_str = str(e).lower()
        if 'encrypt' in error_str or 'password' in error_str:
            file_info['is_encrypted'] = True

            return False, "Encrypted Excel files are not allowed", None

        return False, f"Invalid Excel file: {str(e)}", None

    except Exception as unexpected_error:
        logger.error(
            "Unexpected error reading Excel file: %s",
            str(unexpected_error), exc_info=True
        )

        return False, f"Failed to read Excel file: {str(unexpected_error)}", None

    finally:
        if vba_parser:
            try:
                vba_parser.close()
            except (AttributeError, RuntimeError, IOError) as e:
                logger.debug("Error closing vba_parser: %s", e)

def validate_excel_content(file_path: Path) -> Tuple[bool, str, Optional[Dict]]:
    """
    Main entry point for Excel file content validation.
    
    Orchestrates the validation process based on file extension:
    - .xlsx files: Routes to _validate_xlsx
    - .xls files: Routes to _validate_xls
    
    Also calculates SHA-256 hash of the file for integrity verification.
    
    Args:
        file_path (Path): Path to the Excel file to validate
        
    Returns:
        Tuple[bool, str, Optional[Dict]]: 
            - bool: True if validation passed, False otherwise
            - str: Error message if validation failed, empty string otherwise
            - Optional[Dict]: File metadata including sheets, rows, columns, hash
            
    Example:
        >>> is_valid, error, info = validate_excel_content(Path("data.xlsx"))
        >>> if is_valid:
        ...     print(f"Valid file with {info['total_rows']} rows")
    """
    try:
        file_info = {
            'sheets': [],
            'total_rows': 0,
            'total_cols': 0,
            'has_macros': False,
            'is_encrypted': False,
            'file_hash': None,
            'format': 'unknown'
        }

        # Calculate file hash for integrity checking
        sha256_hash = hashlib.sha256()
        with open(file_path, 'rb') as f:
            for byte_block in iter(lambda: f.read(4096), b''):
                sha256_hash.update(byte_block)

        file_info['file_hash'] = sha256_hash.hexdigest()

        # Validate based on file extension
        if file_path.suffix.lower() == '.xlsx':
            file_info['format'] = 'xlsx'
            return _validate_xlsx(file_path, file_info)

        elif file_path.suffix.lower() == '.xls':
            file_info['format'] = 'xls'
            return _validate_xls(file_path, file_info)

        else:
            return False, f"Unsupported file format: {file_path.suffix}", None

    except MemoryError:
        return False, "File too large to process in memory", None

    except Exception as unexpected_error:
        logger.error("Excel validation error: %s", str(unexpected_error), exc_info=True)
        return False, f"Excel validation failed: {str(unexpected_error)}", None

def validate_file(file) -> Tuple[bool, str, Optional[Dict], int]:
    """
    Perform basic file validation before content processing.
    
    Checks:
    - File existence and non-empty filename
    - File extension against allowed Excel formats
    - File size against MAX_FILE_SIZE limit
    - Content-Length header consistency
    
    Args:
        file: File object from Flask request.files
        
    Returns:
        Tuple[bool, str, Optional[Dict], int]: 
            - bool: True if validation passed, False otherwise
            - str: Error message if validation failed, empty string otherwise
            - Optional[Dict]: Basic file info (extension, original filename)
            - int: HTTP status code appropriate for the result
            
    HTTP Status Codes:
        - 200: Validation passed
        - 400: Bad request (no file, invalid extension)
        - 413: File too large
    """
    if not file:
        return False, "No file provided", None, 400

    if file.filename is None or file.filename == '':
        return False, "No file selected", None, 400

    # Check extension
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ALLOWED_EXCEL_EXTENSIONS:
        return (
            False,
            f"Only Excel files are allowed: {', '.join(ALLOWED_EXCEL_EXTENSIONS)}",
            None, 400
        )

    # Check size
    file_content_length = None
    request_content_length = None

    if hasattr(file, 'content_length') and file.content_length:
        file_content_length = file.content_length

    if hasattr(request, 'content_length') and request.content_length:
        request_content_length = request.content_length

    if file_content_length and request_content_length:
        if file_content_length != request_content_length:
            logger.warning(
                "Size mismatch: file.content_length=%d vs request.content_length=%d",
                file_content_length, request_content_length
            )
            # Use the larger value for safety
            content_length = max(file_content_length, request_content_length)
        else:
            content_length = file_content_length
            logger.debug("Size verified: %d bytes from both sources", content_length)

    # Use whichever is available
    elif file_content_length:
        content_length = file_content_length
        logger.debug("Using file.content_length: %d bytes", content_length)

    elif request_content_length:
        content_length = request_content_length
        logger.debug("Using request.content_length: %d bytes", content_length)

    else:
        logger.warning("No content_length information available")
        return True, "", {'extension': file_ext, 'original_filename': file.filename}, 200

    if content_length > MAX_FILE_SIZE:
        return False, (
            f"File too large. Maximum size: {MAX_FILE_SIZE_MB} MB, "
            f"The size of provided file: {round(content_length / (1024 * 1024), 2)} MB"
        ), None, 413

    return True, "", {'extension': file_ext, 'original_filename': file.filename}, 200

def generate_unique_filename(original_filename: str) -> tuple[str, str, str]:
    """
    Generate a unique, sanitized filename with timestamp and UUID.
    
    Creates a filename in format: {sanitized_base}_{timestamp}_{unique_id}{extension}
    
    Args:
        original_filename (str): Original filename from upload
        
    Returns:
        tuple[str, str, str]: 
            - str: Safe unique filename
            - str: Timestamp string (YYYYMMDD_HHMMSS format)
            - str: Unique ID (first 8 chars of UUID4 hex)
            
    Example:
        >>> safe_name, ts, uid = generate_unique_filename("My File@2.xlsx")
        >>> print(safe_name)  # My File_2_20260226_143015_a1b2c3d4.xlsx
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_id = uuid.uuid4().hex[:8]

    file_ext = Path(original_filename).suffix
    base_name = Path(original_filename).stem

    # Sanitize filename
    safe_base = "".join(c for c in base_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
    if not safe_base:
        safe_base = "file"

    safe_filename = f"{safe_base}_{timestamp}_{unique_id}{file_ext}"

    return safe_filename, timestamp, unique_id

def save_file_with_backup(file_content: bytes, file_path: Path) -> bool:
    """
    Save file with automatic backup creation if file already exists.
    
    If the target file exists, creates a backup copy in BACKUP_DIR with
    timestamp suffix before overwriting.
    
    Args:
        file_content (bytes): File content as bytes
        file_path (Path): Destination path for the file
        
    Returns:
        bool: True if save successful, False otherwise
        
    Note:
        Backup filename format: {original_name}.{timestamp}.bak
    """
    try:
        if file_path.exists():
            backup_path = BACKUP_DIR / f"{file_path.name}.{datetime.now().strftime('%Y%m%d_%H%M%S')}.bak"
            shutil.copy2(file_path, backup_path)
            logger.info("Backup created: %s", backup_path)

        file_path.write_bytes(file_content)
        logger.info("File saved: %s (%d bytes)", file_path, len(file_content))

        return True

    except (IOError, OSError, PermissionError) as e:
        logger.error("Failed to save file %s: %s", file_path, e)
        return False

def trigger_airflow_dag(file_info: Dict[str, Any]) -> Tuple[bool, Optional[Dict]]:
    """
    Trigger Airflow DAG with file information via REST API.
    
    Sends a POST request to Airflow API to trigger the configured DAG
    with file metadata passed in the DAG run configuration.
    
    Args:
        file_info (Dict[str, Any]): Complete file metadata including:
            - safe_filename: Unique filename
            - file_path: Full path to saved file
            - timestamp: Upload timestamp
            - unique_id: File identifier
            - original_filename: Original uploaded name
            - file_hash: SHA-256 hash
            - file_size: Size in bytes
            - sheets: List of sheet information
            - total_rows: Total row count
            - file_format: Excel format (xlsx/xls)
            
    Returns:
        Tuple[bool, Optional[Dict]]:
            - bool: True if DAG triggered successfully
            - Optional[Dict]: DAG run response on success, error dict on failure
            
    Raises:
        requests.exceptions.RequestException: For network/connection errors
    """
    try:
        auth = (AIRFLOW_USER, AIRFLOW_PASSWORD)
        dag_run_url = f"{AIRFLOW_API_URL}/dags/{DAG_ID}/dagRuns"

        payload = {
            "conf": {
                "file_name": file_info['safe_filename'],
                "file_path": str(file_info['final_path']),
                "upload_timestamp": file_info['timestamp'],
                "unique_id": file_info['unique_id'],
                "original_filename": file_info['original_filename'],
                "file_hash": file_info.get('file_hash'),
                "file_size": file_info.get('file_size'),
                "sheets": file_info.get('sheets', []),
                "total_rows": file_info.get('total_rows', 0),
                "file_format": file_info.get('format', 'unknown')
            }
        }

        logger.info("Triggering DAG %s with file: %s", DAG_ID, file_info['safe_filename'])

        response = requests.post(
            dag_run_url,
            json=payload,
            auth=auth,
            timeout=30,
            headers={'Content-Type': 'application/json'}
        )

        if response.status_code in (200, 201):
            logger.info("DAG triggered successfully")
            return True, response.json()
        else:
            logger.error("Failed to trigger DAG. Status: %s", response.status_code)
            return False, {'status_code': response.status_code, 'error': response.text}

    except requests.exceptions.ConnectionError as e:
        logger.error("Connection error to Airflow API: %s", e)
        return False, {'error': 'Airflow API connection failed'}

    except requests.exceptions.Timeout as e:
        logger.error("Timeout connecting to Airflow API: %s", e)
        return False, {'error': 'Airflow API timeout'}

    except requests.exceptions.RequestException as e:
        logger.error("Request error to Airflow API: %s", e)
        return False, {'error': f'Airflow API error: {str(e)}'}

    except Exception as unexpected_error:
        logger.error("Unexpected error triggering DAG: %s", unexpected_error, exc_info=True)
        return False, {'error': f'Unexpected error: {str(unexpected_error)}'}

def cleanup_old_files(target_dir: Path, hours: int) -> int:
    """
    Delete files in target directory older than specified hours.
    
    Uses file modification time (st_mtime) to determine age.
    
    Args:
        target_dir (Path): Directory to clean up
        hours (int): Age threshold in hours
        
    Returns:
        int: Number of files successfully deleted
        
    Note:
        Errors during deletion are logged but don't stop the cleanup process
    """
    try:
        cutoff_time = datetime.now() - timedelta(hours=hours)
        deleted = 0

        for file_path in target_dir.glob('*'):
            if file_path.is_file():
                try:
                    if datetime.fromtimestamp(file_path.stat().st_mtime) < cutoff_time:
                        file_path.unlink()
                        deleted += 1
                        logger.debug("Deleted old file: %s", file_path)

                except (OSError, PermissionError) as e:
                    logger.warning("Failed to delete %s: %s", file_path, e)

        if deleted > 0:
            logger.info("Cleaned up %d old files from %s", deleted, target_dir)

        return deleted

    except Exception as unexpected_error:
        logger.error("Cleanup error in %s: %s", target_dir, unexpected_error)
        return 0

# ========== RATE LIMITING DECORATOR ==========
def rate_limit(limit_string: Optional[str] = None):
    """
    Decorator factory for applying rate limits to endpoints.
    
    Wraps Flask routes with Flask-Limiter's rate limiting functionality.
    
    Args:
        limit_string (Optional[str]): Rate limit string (e.g., "10 per minute").
                                     If None, uses default RATE_LIMIT setting.
                                     
    Returns:
        Callable: Decorated function with rate limiting applied
        
    Example:
        >>> @rate_limit("5 per minute")
        ... def my_endpoint():
        ...     return jsonify({"message": "ok"})
    """
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            return limiter.limit(limit_string or RATE_LIMIT)(f)(*args, **kwargs)
        return wrapped
    return decorator

# ========== API ENDPOINTS ==========
@upload_bp.route('/health', methods=['GET'])
def health_check():
    """
    Health check endpoint for monitoring and container orchestration.
    
    Returns:
        Response: JSON with:
            - status: "healthy" if service is operational
            - timestamp: Current ISO timestamp
            - clamav_available: ClamAV service status
            - upload_dir_exists: Upload directory existence
            - upload_dir_writable: Write permissions status
            
    HTTP Status:
        - 200: Service is healthy
    """
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'clamav_available': clamav_scanner.is_available(),
        'upload_dir_exists': UPLOAD_DIR.exists(),
        'upload_dir_writable': os.access(UPLOAD_DIR, os.W_OK) if UPLOAD_DIR.exists() else False
    }), 200

@upload_bp.route('/stats', methods=['GET'])
def get_stats():
    """
    Get detailed upload statistics (admin only endpoint).
    
    In production mode, only allows access from localhost.
    
    Returns:
        Response: JSON with:
            - clamav: ClamAV statistics
            - uploads: Directory status and file count
            - retention_hours: Current retention setting
            - max_file_size_mb: Maximum allowed file size
            - rate_limit: Current rate limit setting
            
    HTTP Status:
        - 200: Success
        - 403: Forbidden (production and non-localhost)
    """
    if IS_PRODUCTION:
        if request.remote_addr not in ['127.0.0.1', '::1']:
            return jsonify({'error': 'Forbidden'}), 403

    stats = {
        'clamav': clamav_scanner.get_stats(),
        'uploads': {
            'directory': str(UPLOAD_DIR),
            'exists': UPLOAD_DIR.exists(),
            'writable': os.access(UPLOAD_DIR, os.W_OK) if UPLOAD_DIR.exists() else False
        },
        'retention_hours': FILE_RETENTION_DAYS * 24,
        'max_file_size_mb': MAX_FILE_SIZE_MB,
        'rate_limit': RATE_LIMIT
    }

    try:
        if UPLOAD_DIR.exists():
            files = list(UPLOAD_DIR.glob('*'))
            stats['uploads']['file_count'] = len([f for f in files if f.is_file()])

    except Exception as e:
        stats['uploads']['error'] = str(e)

    return jsonify(stats), 200

@upload_bp.route('/cleanup', methods=['POST'])
def manual_cleanup():
    """
    Manually trigger cleanup of old files (admin only endpoint).
    
    In production mode, only allows access from localhost.
    
    Request Body:
        JSON with optional 'hours' parameter (defaults to retention setting)
        
    Returns:
        Response: JSON with:
            - message: "Cleanup completed"
            - deleted: Counts of deleted items by category
            - hours: Hours threshold used
            
    HTTP Status:
        - 200: Cleanup completed
        - 403: Forbidden (production and non-localhost)
    """
    if IS_PRODUCTION:
        if request.remote_addr not in ['127.0.0.1', '::1']:
            return jsonify({'error': 'Forbidden'}), 403

    hours = request.json.get('hours', FILE_RETENTION_DAYS * 24) if request.json else FILE_RETENTION_DAYS * 24

    results = {
        'uploads': cleanup_old_files(UPLOAD_DIR, hours),
        'backups': cleanup_old_files(BACKUP_DIR, hours),
        'temp': clamav_scanner.manual_cleanup(max_age_hours=hours) if clamav_scanner.is_available() else 0,
        'metadata': file_metadata_store.cleanup_expired()
    }

    return jsonify({
        'message': 'Cleanup completed',
        'deleted': results,
        'hours': hours
    }), 200

@upload_bp.route('/upload-mft-excel', methods=['POST'])
@rate_limit()
def upload_mft_excel():
    """
    Main endpoint for secure Excel file upload with virus scanning and validation.
    
    Comprehensive upload process:
    1. Basic validation (file presence, size, extension)
    2. Read file into memory
    3. Virus scan via ClamAV (if available)
    4. Generate unique filename
    5. Save temporary file
    6. Excel content validation (macros, encryption, structure)
    7. Move to permanent location
    8. Store metadata
    9. Trigger Airflow DAG
    
    Returns:
        Response: JSON with:
            - message: Success/failure message
            - file: File metadata
            - dag_triggered: Boolean indicating DAG trigger status
            - warning/dag_error: Optional warnings if DAG trigger fails
            
    HTTP Status:
        - 200: Complete success
        - 207: Partial success (file saved but DAG not triggered)
        - 400: Validation error or virus detected
        - 413: File too large
        - 500: Server error
        
    Rate Limited: Yes (configured by RATE_LIMIT)
    """
    temp_path = None
    start_time = time.time()

    try:
        session['last_upload_attempt'] = datetime.now().isoformat()

        # 1. Check file existence
        if 'file' not in request.files:
            logger.error("No file provided in request")
            return jsonify({'error': 'No file provided'}), 400

        file = request.files['file']

        if file.filename is None:
            logger.error("Filename is None, cannot scan")
            return jsonify({'error': 'Invalid filename'}), 400

        # 2. Basic validation
        is_valid, error_msg, file_info, http_status  = validate_file(file)
        if not is_valid:
            logger.error("Validation failed: %s", error_msg)
            return jsonify({'error': error_msg}), http_status

        if file_info is None:
            file_info = {}
            logger.debug("file_info was None, initialized to empty dict")

        logger.info("File received: %s", file.filename)

        # 3. Read file into memory
        file_content = file.read()
        file_size = len(file_content)

        logger.info("File read into memory: %d bytes", file_size)

        # 4. Virus scan
        if clamav_scanner.is_available():
            logger.info("Starting virus scan for: %s", file.filename)

            is_clean, scan_result = clamav_scanner.scan_stream(file_content, file.filename)

            if not is_clean:
                logger.error("VIRUS DETECTED: %s", scan_result)

                # Save to quarantine
                quarantine_filename = f"INFECTED_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}_{file.filename}"
                quarantine_path = QUARANTINE_DIR / quarantine_filename

                try:
                    quarantine_path.write_bytes(file_content)
                    logger.info("Infected file saved to quarantine: %s", quarantine_path)

                except (IOError, OSError, PermissionError) as e:
                    logger.error(
                        "File system error while saving infected file to quarantine: %s", e
                    )

                    # Try to save infected file to a temporary directory
                    try:
                        fallback_path = Path(tempfile.gettempdir()) / quarantine_filename
                        fallback_path.write_bytes(file_content)
                        logger.warning(
                            "Infected file saved to fallback location: %s", fallback_path
                        )
                        quarantine_path = fallback_path

                    except (IOError, OSError, PermissionError, MemoryError) as fallback_error:
                        logger.error("Fallback save also failed: %s", fallback_error)
                        quarantine_path = None

                except (ValueError, TypeError) as e:
                    logger.error("Invalid parameter for quarantine file: %s", e)
                    quarantine_path = None

                except MemoryError as e:
                    logger.error("Memory error while saving infected file: %s", e)
                    quarantine_path = None

                except Exception as unexpected_error:
                    logger.error(
                        "Unexpected error saving infected file: %s",
                        unexpected_error, exc_info=True
                    )
                    quarantine_path = None

                return jsonify({
                    'error': 'File rejected: virus detected',
                    'virus': scan_result,
                    'quarantine_id': quarantine_filename if quarantine_path and quarantine_path.exists() else None
                }), 400

            logger.info("Virus scan passed")

        else:
            logger.warning("ClamAV not available - skipping virus scan")

        # 5. Generate unique filename
        safe_filename, timestamp, unique_id = generate_unique_filename(file.filename)
        final_file_path = UPLOAD_DIR / safe_filename

        # 6. Save temporary file for validation
        temp_path = UPLOAD_DIR / f".tmp_{safe_filename}"
        temp_path.write_bytes(file_content)
        logger.info("Temporary file created: %s", temp_path)

        # 7. Validate Excel content
        is_valid_content, content_error, content_info = validate_excel_content(temp_path)
        if not is_valid_content:
            logger.error("Content validation failed: %s", content_error)
            temp_path.unlink(missing_ok=True)
            return jsonify({'error': content_error}), 400

        # Merge content info
        if content_info:
            file_info.update(content_info)

        # 8. Move to final location
        shutil.move(str(temp_path), str(final_file_path))
        temp_path = None
        logger.info("File saved to: %s", final_file_path)

        # 9. Prepare metadata
        file_metadata = {
            'safe_filename': safe_filename,
            'final_path': final_file_path,
            'timestamp': timestamp,
            'unique_id': unique_id,
            'original_filename': file.filename,
            'file_size': file_size,
            'file_hash': file_info.get('file_hash'),
            'sheets': file_info.get('sheets', []),
            'total_rows': file_info.get('total_rows', 0),
            'has_macros': file_info.get('has_macros', False),
            'format': file_info.get('format', 'unknown'),
            'upload_duration_ms': int((time.time() - start_time) * 1000)
        }

        file_metadata_store.add(unique_id, file_metadata)

        # 10. Trigger DAG
        dag_success, dag_response = trigger_airflow_dag(file_metadata)

        # 11. Response
        response_data = {
            'message': 'File uploaded successfully',
            'file': {
                'name': safe_filename,
                'original_name': file.filename,
                'size': file_size,
                'upload_time': timestamp,
                'id': unique_id,
                'sheets': file_info.get('sheets', []),
                'total_rows': file_info.get('total_rows', 0),
                'format': file_info.get('format', 'unknown'),
                'hash': file_info.get('file_hash')
            },
            'dag_triggered': dag_success
        }

        if dag_success:
            response_data['dag'] = dag_response
            return jsonify(response_data), 200
        else:
            response_data['warning'] = 'File saved but DAG trigger failed'
            response_data['dag_error'] = dag_response
            return jsonify(response_data), 207

    except MemoryError:
        return jsonify({'error': 'File too large to process'}), 413

    except (IOError, OSError, PermissionError) as e:
        logger.error("File system error: %s", e, exc_info=True)
        return jsonify({'error': f'File system error: {str(e)}'}), 500

    except Exception as unexpected_error:
        logger.error("Unexpected error: %s", unexpected_error, exc_info=True)
        return jsonify({'error': f'Upload failed: {str(unexpected_error)}'}), 500

    finally:
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
                logger.debug("Temporary file cleaned up: %s", temp_path)

            except (OSError, PermissionError, FileNotFoundError) as e:
                logger.debug("Could not delete temp file %s (normal): %s", temp_path, e)

            except Exception as unexpected_error:
                logger.error(
                    "Unexpected error while cleaning up temp file %s: %s", 
                    temp_path, unexpected_error, exc_info=True
                )

@upload_bp.route('/file/<file_id>', methods=['GET'])
def get_file_info(file_id: str):
    """
    Retrieve metadata for a previously uploaded file by ID.
    
    Args:
        file_id (str): Unique identifier returned from upload
        
    Returns:
        Response: JSON with file metadata if found
        
    HTTP Status:
        - 200: File found, metadata returned
        - 404: File ID not found or expired
    """
    metadata = file_metadata_store.get(file_id)
    if metadata:
        return jsonify({'file': metadata}), 200
    return jsonify({'error': 'File not found or expired'}), 404

@upload_bp.route('/file/<file_id>/status', methods=['GET'])
def get_file_status(file_id: str):
    """
    Check if an uploaded file still exists in the filesystem.
    
    Args:
        file_id (str): Unique identifier returned from upload
        
    Returns:
        Response: JSON with:
            - file_id: Requested file ID
            - exists: Boolean indicating file existence
            - path: Full path if file exists
            - metadata: Complete metadata if file exists
            
    HTTP Status:
        - 200: Check completed (file may or may not exist)
        - 404: File ID not found in metadata store
    """
    metadata = file_metadata_store.get(file_id)
    if not metadata:
        return jsonify({'error': 'File not found'}), 404

    file_path = Path(metadata['final_path'])
    exists = file_path.exists()

    return jsonify({
        'file_id': file_id,
        'exists': exists,
        'path': str(file_path) if exists else None,
        'metadata': metadata if exists else None
    }), 200

# ========== BACKGROUND CLEANUP TASK ==========
def start_background_cleanup():
    """
    Start background daemon thread for periodic file cleanup.
    
    Creates a thread that runs every hour and performs:
    - Cleanup of old upload files
    - Cleanup of old backup files
    - Expired metadata removal
    - ClamAV temporary file cleanup
    
    The thread continues running until the main process terminates.
    
    Note:
        Errors in one cleanup category don't affect others
    """
    def cleanup_worker():
        while True:
            try:
                time.sleep(3600)  # Run every hour

                hours = FILE_RETENTION_DAYS * 24
                deleted_uploads = deleted_backups = expired_metadata = 0

                # Cleanup with error isolation
                try:
                    deleted_uploads = cleanup_old_files(UPLOAD_DIR, hours)
                except (OSError, PermissionError, FileNotFoundError) as e:
                    logger.error("Uploads cleanup failed: %s", e)
                except Exception as unexpected_error:
                    logger.error("Unexpected uploads error: %s", unexpected_error, exc_info=True)

                try:
                    deleted_backups = cleanup_old_files(BACKUP_DIR, hours)
                except (OSError, PermissionError, FileNotFoundError) as e:
                    logger.error("Backups cleanup failed: %s", e)
                except Exception as unexpected_error:
                    logger.error("Unexpected backups error: %s", unexpected_error, exc_info=True)

                try:
                    expired_metadata = file_metadata_store.cleanup_expired()
                except (KeyError, ValueError, AttributeError) as e:
                    logger.error("Metadata cleanup failed: %s", e)
                except Exception as unexpected_error:
                    logger.error("Unexpected metadata error: %s", unexpected_error, exc_info=True)

                if any([deleted_uploads, deleted_backups, expired_metadata]):
                    logger.info(
                        "Background cleanup: %d uploads, %d backups, %d metadata",
                        deleted_uploads, deleted_backups, expired_metadata
                    )

            except (KeyboardInterrupt, SystemExit):
                logger.info("Background cleanup thread stopped")
                break

            except MemoryError as e:
                logger.error("Memory error in cleanup: %s", e)
                time.sleep(60)

            except Exception as unexpected_error:
                logger.error("Critical cleanup error: %s", unexpected_error, exc_info=True)
                time.sleep(60)

    thread = threading.Thread(target=cleanup_worker, daemon=True)
    thread.start()
    logger.info("Background cleanup thread started")

# ========== FLASK APP SETUP ==========
def create_app():
    """
    Create and configure the Flask application instance.
    
    Sets up:
    - Secret key for sessions
    - Maximum content length
    - Upload folder configuration
    - Blueprint registration
    - Rate limiter initialization
    - Background cleanup thread
    
    Returns:
        Flask: Configured Flask application instance
    """
    app = Flask(__name__)
    app.secret_key = FLASK_SECRET_KEY
    app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE
    app.config['UPLOAD_FOLDER'] = str(UPLOAD_DIR)

    app.register_blueprint(upload_bp)
    limiter.init_app(app)
    start_background_cleanup()

    return app

app = create_app()

if __name__ == '__main__':
    logger.info("Starting Upload API on %s:%s", FLASK_HOST, FLASK_PORT)
    logger.info("Environment: %s", FLASK_ENV)
    logger.info("Max file size: %d MB", MAX_FILE_SIZE_MB)
    logger.info("Rate limit: %s", RATE_LIMIT)

    app.run(
        host=FLASK_HOST,
        port=FLASK_PORT,
        debug=FLASK_DEBUG,
        threaded=True
    )
